from celery import shared_task
import os
import openpyxl
from PIL import Image, ImageDraw
from fitz import fitz
import aspose.words as aw

from notifications.models import NotificationModel
from users.models import CustomUser
from .models import FileModel, ChatMessage, ChatRoom, Message
from django.db import models


@shared_task
def create_message_pdf_image(file_path, file_id, chat_message_id,model_name:models.Model, dpi=150):
    doc = fitz.open(file_path)

    file_name = file_path.split('/')[-1].split('.')[0]

    # Select the first page
    page = doc.load_page(0)  # 0 is the index for the first page

    # Set the zoom factor for the image. Higher values will result in higher resolution.
    zoom = dpi / 72  # PDF standard DPI is 72
    mat = fitz.Matrix(zoom, zoom)
    output_image_path = f"/my_future/media/message_image/{file_name}.jpg"
    if os.path.exists(output_image_path):
        output_image_path = f"/my_future/media/message_image/{file_name.split('.')[0]}_{file_name}.jpg"
    # Render the page to an image (pixmap)
    pix = page.get_pixmap(matrix=mat)

    # Save the image as a JPG file
    pix.save(output_image_path)
    try:
        file_get = FileModel.objects.get(id=file_id)
        file_get.file_image = output_image_path
        file_get.save()
        chat_message = model_name.objects.get(id=chat_message_id)
        chat_message.save()
    except:
        pass


@shared_task
def create_message_excel_image(file_path, file_id, chat_message_id, model_name: models.Model):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    file_name = file_path.split('/')[-1].split('.')[0]
    sheet = wb.worksheets[0]
    img = Image.new('RGB', (sheet.max_column * 100, sheet.max_row * 20), 'white')
    d = ImageDraw.Draw(img)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                # Calculate position based on cell's row and column
                x = (cell.column - 1) * 100
                y = (cell.row - 1) * 20
                message = str(cell.value)
                # Draw the cell value
                # If you've loaded a font, add `, font=font` after `fill="black"`
                d.text((x, y), message, fill="black")

    output_image_path = f"/my_future/media/message_image/{file_name}.jpg"
    if os.path.exists(output_image_path):
        output_image_path = f"/my_future/media/message_image/{file_name.split('.')[0]}_{file_name}.jpg"
    img.save(output_image_path)
    try:
        file_get = FileModel.objects.get(id=file_id)
        file_get.file_image = output_image_path
        file_get.save()
        chat_message = model_name.objects.get(id=chat_message_id)
        chat_message.save()
    except:
        pass


@shared_task
def forward_post_task(post, creator, forward_users_id):

    forward_users = CustomUser.objects.filter(id__in=forward_users_id)
    # Check if the mentioned users are in the chat room
    for receiver in forward_users:
        check_chat_room = ChatRoom.objects.filter(first_user=receiver, second_user=creator)
        second_check_chat_room = ChatRoom.objects.filter(first_user=creator, second_user=receiver)
        # If the mentioned users are not in the chat room, create a chat room and send the story
        if creator != receiver and not check_chat_room.exists() and not second_check_chat_room.exists():
            chat_room = ChatRoom.objects.create(first_user=creator, second_user=receiver)
            message = Message.objects.create(message_type='forward_post', forward_post=post)
            ChatMessage.objects.create(chat_room=chat_room, message=message, from_user=creator)
            # Create a notification for the mentioned user
            NotificationModel.objects.create(
                forward_post=post,
                notification_visible_to_user=receiver,
                type_notification=6)

        else:
            # If the mentioned users are in the chat room, send the story
            if check_chat_room.exists():
                message = Message.objects.create(message_type='forward_post', forward_post=post)
                ChatMessage.objects.create(chat_room=check_chat_room.last(), message=message, from_user=creator)
                NotificationModel.objects.create(
                    forward_post=post,
                    notification_visible_to_user=receiver,
                    type_notification=6)
            # If the mentioned users are in the chat room, send the story
            elif second_check_chat_room.exists():
                message = Message.objects.create(message_type='forward_post', forward_post=post)
                ChatMessage.objects.create(chat_room=second_check_chat_room.last(), message=message, from_user=creator)
                NotificationModel.objects.create(
                    forward_post=post,
                    notification_visible_to_user=receiver,
                    type_notification=6)
            else:
                pass
